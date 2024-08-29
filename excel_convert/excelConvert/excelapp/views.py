from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import UploadFileForm, ConfigurationFileForm
from .models import ConfigurationFile
import openpyxl
import os
import json
import csv
from django.conf import settings
import logging
import math

logger = logging.getLogger(__name__)

def config_upload(request):
    if request.method == 'POST':
        form = ConfigurationFileForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('config_list')
    else:
        form = ConfigurationFileForm()
    return render(request, 'excelapp/config_upload.html', {'form': form})

def config_list(request):
    configurations = ConfigurationFile.objects.all()
    return render(request, 'excelapp/config_list.html', {'configurations': configurations})

def config_edit(request, pk):
    config = get_object_or_404(ConfigurationFile, pk=pk)
    if request.method == 'POST':
        form = ConfigurationFileForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            return redirect('config_list')
    else:
        form = ConfigurationFileForm(instance=config)
    return render(request, 'excelapp/config_edit.html', {'form': form, 'config': config})

def config_delete(request, pk):
    config = get_object_or_404(ConfigurationFile, pk=pk)
    if request.method == 'POST':
        config.delete()
        return redirect('config_list')
    return render(request, 'excelapp/config_delete_confirm.html', {'config': config})

def format_conversion(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            config_id = request.POST.get('config_id')
            
            config = get_object_or_404(ConfigurationFile, id=config_id)
            column_mapping = config.content

            file_path = os.path.join(settings.MEDIA_ROOT, file.name)
            
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            
            wb = openpyxl.load_workbook(file_path)
            
            logger.info(f"Column mapping: {column_mapping}")

            sheet_data = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    data.append(row_data)
                sheet_data[sheet_name] = {'headers': headers, 'data': data}
            
            logger.info(f"Sheet data: {sheet_data}")

            key_column = next((k for k, v in column_mapping.items() if v.get('keyFlag') == 'true'), None)
            if key_column:
                key_data = column_mapping[key_column]['data']
                logger.info(f"Key column: {key_column}, Key data: {key_data}")

                if len(wb.sheetnames) >= 2:
                    sheet1_data = sheet_data[wb.sheetnames[0]]['data']
                    sheet2_data = sheet_data[wb.sheetnames[1]]['data']
                    
                    sheet1_dict = {row[key_data]: row for row in sheet1_data if key_data in row and row[key_data] is not None and str(row[key_data]).strip() != ""}
                    sheet2_dict = {row[key_data]: row for row in sheet2_data if key_data in row and row[key_data] is not None and str(row[key_data]).strip() != ""}
                    
                    for key, row in sheet1_dict.items():
                        if key in sheet2_dict:
                            row.update(sheet2_dict[key])
                    
                    combined_data = list(sheet1_dict.values())
                    sheet_data['combined'] = {'headers': sheet_data[wb.sheetnames[0]]['headers'] + sheet_data[wb.sheetnames[1]]['headers'], 'data': combined_data}
                else:
                    sheet_data['combined'] = {
                        'headers': sheet_data[wb.sheetnames[0]]['headers'],
                        'data': [row for row in sheet_data[wb.sheetnames[0]]['data'] if key_data in row and row[key_data] is not None and str(row[key_data]).strip() != ""]
                    }
            else:
                logger.warning("No key column specified. Using first sheet data as combined data.")
                sheet_data['combined'] = sheet_data[wb.sheetnames[0]]

            logger.info(f"Combined data: {sheet_data['combined']}")

            new_data = []
            new_headers = list(column_mapping.keys())
            new_data.append(new_headers)

            for row in sheet_data['combined']['data']:
                new_row = []
                for output_header, mapping in column_mapping.items():
                    logger.info(f"Processing: {output_header} -> {mapping}")
                    if mapping.get('kotei'):
                        new_row.append(mapping['kotei'])
                    else:
                        try:
                            column_name = mapping['data']
                            value = row.get(column_name)
                            logger.info(f"Original value for {column_name}: {value}")
                            
                            if value is not None and str(value).strip() != "":
                                try:
                                    numeric_value = float(value)
                                    if math.isnan(numeric_value):
                                        value = ""
                                    else:
                                        if mapping.get('numMinus'):
                                            num_minus = float(mapping['numMinus'])
                                            numeric_value = max(0, numeric_value - num_minus)
                                        if mapping.get('minusFlag') == 'true':
                                            numeric_value = -numeric_value
                                        value = str(numeric_value) if numeric_value != 0 else ""
                                except ValueError:
                                    if mapping.get('minusFlag') == 'true':
                                        value = f"-{value}"
                            else:
                                value = ""
                            logger.info(f"Final value for {column_name}: {value}")
                            new_row.append(value)
                        except Exception as e:
                            logger.error(f"Error processing {output_header}: {str(e)}")
                            new_row.append("")
                new_data.append(new_row)

            response = HttpResponse(content_type='text/csv; charset=shift_jis')
            response['Content-Disposition'] = f'attachment; filename="modified_{os.path.splitext(file.name)[0]}.csv"'

            writer = csv.writer(response)
            for row in new_data:
                encoded_row = [str(cell).encode('shift_jis', 'ignore').decode('shift_jis') for cell in row]
                writer.writerow(encoded_row)

            return response
    else:
        form = UploadFileForm()
        configurations = ConfigurationFile.objects.all()
    return render(request, 'excelapp/upload.html', {'form': form, 'configurations': configurations})