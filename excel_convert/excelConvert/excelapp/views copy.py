from django.shortcuts import render
from django.http import HttpResponse
from .forms import UploadFileForm
import openpyxl
import os
import json
import csv
from django.conf import settings
import logging

logger = logging.getLogger(__name__)
def config_upload(request):
    # ここに設定ファイル登録の処理を実装します
    return render(request, 'excelapp/config_upload.html')

def format_conversion(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_path = os.path.join(settings.MEDIA_ROOT, file.name)
            
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            
            wb = openpyxl.load_workbook(file_path)
            
            # 設定ファイルを読み込む（エンコーディングを指定）
            with open(os.path.join(settings.BASE_DIR, 'excelapp/itemMapping/フリーウェイにインポートする.json'), 'r', encoding='utf-8') as f:
                column_mapping = json.load(f)
            
            logger.info(f"Column mapping: {column_mapping}")

            # シートごとのデータを保持する辞書
            sheet_data = {}

            # キー項目を特定
            key_column = next((k.split('_')[1][:-3] for k in column_mapping.keys() if k.endswith('(*)')), None)
            logger.info(f"Key column: {key_column}")

            # 各シートのデータを辞書に格納
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = {cell.value: cell.column_letter for cell in ws[1]}
                data = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    row_data = {col: row[openpyxl.utils.column_index_from_string(headers[col]) - 1].value for col in headers}
                    data.append(row_data)
                sheet_data[sheet_name] = {'headers': headers, 'data': data}
            
            logger.info(f"Sheet data: {sheet_data}")

            # シート1とシート2を結合
            if len(wb.sheetnames) >= 2:
                sheet1_data = sheet_data[wb.sheetnames[0]]['data']
                sheet2_data = sheet_data[wb.sheetnames[1]]['data']
                
                # キー項目でデータを辞書化
                sheet1_dict = {row[key_column]: row for row in sheet1_data}
                sheet2_dict = {row[key_column]: row for row in sheet2_data}
                
                # シート1とシート2を結合
                for key, row in sheet1_dict.items():
                    if key in sheet2_dict:
                        row.update(sheet2_dict[key])
                
                combined_data = list(sheet1_dict.values())
                sheet_data['combined'] = {'headers': {**sheet_data[wb.sheetnames[0]]['headers'], **sheet_data[wb.sheetnames[1]]['headers']}, 'data': combined_data}
            
            logger.info(f"Combined data: {sheet_data['combined']}")

            # 新しいデータを保持するリスト
            new_data = []

            # 新しいヘッダーを設定し、列を並べ替える
            new_headers = [new_header for old_header, new_header in column_mapping.items()]
            new_data.append(new_headers)

            # 2行目以降のデータを入れ替え
            for row in sheet_data['combined']['data']:
                new_row = []
                for old_header, new_header in column_mapping.items():
                    logger.info(f"Processing: {old_header} -> {new_header}")
                    if old_header.startswith("固定値_"):
                        # 固定値の場合
                        fixed_value = old_header.split('_', 1)[1]
                        new_row.append(fixed_value)
                    else:
                        try:
                            column_name = '_'.join(old_header.split('_')[1:])  # すべての '_' 以降を column_name とする
                            
                            # (-) の処理
                            add_minus = False
                            if column_name.endswith("(-)"):
                                add_minus = True
                                column_name = column_name[:-3]  # (-) を除去
                            
                            # (*) の処理
                            if column_name.endswith("(*)"):
                                column_name = column_name[:-3]  # (*) を除去
                            
                            logger.info(f"Column name: {column_name}")
                            
                            if column_name in row:
                                value = row[column_name]
                                logger.info(f"Original value: {value}")
                                if value is not None and str(value).strip() != "":
                                    # 数値に変換を試みる
                                    try:
                                        numeric_value = float(value)
                                        if add_minus:
                                            numeric_value = -numeric_value
                                        value = str(numeric_value)  # 文字列に戻す
                                    except ValueError:
                                        # 数値に変換できない場合は、元の値をそのまま使用
                                        if add_minus:
                                            value = f"-{value}"
                                logger.info(f"Final value: {value}")
                                new_row.append(value)
                            else:
                                logger.warning(f"Column {column_name} not found in row")
                                new_row.append(None)  # デフォルト値を設定
                        except (IndexError, ValueError, KeyError) as e:
                            logger.error(f"Error processing {old_header}: {str(e)}")
                            new_row.append(None)  # デフォルト値を設定
                new_data.append(new_row)

            # CSVファイルとして出力（UTF-8 BOM付き）
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="modified_{os.path.splitext(file.name)[0]}.csv"'
            response.write('\ufeff'.encode('utf8'))  # UTF-8 BOM

            writer = csv.writer(response)
            writer.writerows(new_data)

            return response
    else:
        form = UploadFileForm()
    return render(request, 'excelapp/upload.html', {'form': form})